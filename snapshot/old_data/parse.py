from openpyxl import load_workbook
import json 

wb = load_workbook('D:\\Projects\\edu-prof-app\\edu-prof-app-backend\\data\\x\\SSO1 специальности.xlsx')
ws = wb.active

cellListColumB = ws['B']

SpecialtyListJ = []
SkillListJ = []

curSpecialty = 119
curSkill = 478
for item in cellListColumB:
    # print(item.value)
    if(item.value):
        curSpecialty += 1 
        print("="*24, "")
        print(curSpecialty)
        print("="*24, "")
        print(item.value)
        print(ws.cell(row=item.row, column=1).value)
        print("."*24, "")

        SpecialtyListJ.append({
            "model": "api.Specialty",
                "pk": curSpecialty,
                "fields": {
                    "code": ws.cell(row=item.row, column=1).value,
                    "title": item.value,
                    "c_type":"ПТО"
                }
        })
    else:
        curSkill +=1
        print(ws.cell(row=item.row, column=1).value)
        print(ws.cell(row=item.row, column=3).value)

        SkillListJ.append({
            "model": "api.Skill",
                "pk": curSkill,
                "fields": {
                    "code": ws.cell(row=item.row, column=1).value,
                    "title": ws.cell(row=item.row, column=3).value,
                    "specialty": curSpecialty
                }
        })

print(SpecialtyListJ)

with open("specSSO1.json", "w", encoding='utf-8') as outfile: 
    json.dump(SpecialtyListJ, outfile, indent = 4, ensure_ascii=False)

with open("skillSSO1.json", "w", encoding='utf-8') as outfile: 
    json.dump(SkillListJ, outfile, indent = 4, ensure_ascii=False)