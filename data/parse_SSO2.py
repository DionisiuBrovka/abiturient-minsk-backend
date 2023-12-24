from openpyxl import load_workbook
import json 

wb = load_workbook('D:\\Projects\\edu-prof-app\\edu-prof-app-backend\\data\\x\\SSO2 специальности.xlsx')
ws = wb.active

cellListColumB = ws['B']

SpecialtyListJ = []
SkillListJ = []

curSpecialty = 159
curSkill = 535
for item in cellListColumB:
    # print(item.value)
    if(item.value):
        curSpecialty += 1 
        curSkill +=1
        print("="*24, "")
        print(curSpecialty)
        print("="*24, "")
        print(item.value)
        print(ws.cell(row=item.row, column=1).value)
        print("."*24, "")

        SpecialtyListJ.append({
            "model": "api.Specialty",
                "pk": curSpecialty,
                "fields": {
                    "code": ws.cell(row=item.row, column=1).value,
                    "title": item.value,
                    "c_type":"ПТО"
                }
        })

        SkillListJ.append({
            "model": "api.Skill",
                "pk": curSkill,
                "fields": {
                    "code": ws.cell(row=item.row, column=1).value,
                    "title": ws.cell(row=item.row, column=3).value,
                    "specialty": curSpecialty
                }
        })

        

print(SpecialtyListJ)

with open("specSSO2.json", "w", encoding='utf-8') as outfile: 
    json.dump(SpecialtyListJ, outfile, indent = 4, ensure_ascii=False)

with open("skillSSO2.json", "w", encoding='utf-8') as outfile: 
    json.dump(SkillListJ, outfile, indent = 4, ensure_ascii=False)