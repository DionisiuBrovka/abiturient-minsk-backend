from openpyxl import load_workbook
import json 

wb = load_workbook('D:\\Projects\\edu-prof-app\\edu-prof-app-backend\\data\\x\\PTO специальности.xlsx')
ws = wb.active

cellListColumB = ws['B']

SpecialtyListJ = []
SkillListJ = []

curSpecialty = 0
curSkill = 0
for item in cellListColumB:
    # print(item.value)
    if(item.value):
        curSpecialty += 1 
        print("="*24, "")
        print(curSpecialty)
        print("="*24, "")
        print(item.value)
        print(ws.cell(row=item.row, column=1).value)
        print("."*24, "")

        SpecialtyListJ.append({
            "model": "api.Specialty",
                "pk": curSpecialty,
                "fields": {
                    "code": ws.cell(row=item.row, column=1).value,
                    "title": item.value,
                    "c_type":"ПТО"
                }
        })
    else:
        curSkill +=1
        print(ws.cell(row=item.row, column=1).value)
        print(ws.cell(row=item.row, column=3).value)

        SkillListJ.append({
            "model": "api.Skill",
                "pk": curSkill,
                "fields": {
                    "code": ws.cell(row=item.row, column=1).value,
                    "title": ws.cell(row=item.row, column=3).value,
                    "specialty": curSpecialty
                }
        })

print(SpecialtyListJ)

with open("specPTO.json", "w", encoding='utf-8') as outfile: 
    json.dump(SpecialtyListJ, outfile, indent = 4, ensure_ascii=False)

with open("skillPTO.json", "w", encoding='utf-8') as outfile: 
    json.dump(SkillListJ, outfile, indent = 4, ensure_ascii=False)